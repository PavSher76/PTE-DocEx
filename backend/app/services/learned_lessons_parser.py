from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

LESSON_NUMBER_RE = re.compile(r"^\d+(?:\.\d+)?$")
METADATA_LABELS = {
    "наименование проекта": "project_name",
    "вид документации": "documentation_type",
    "должность": "position",
}

LESSON_FIELD_KEYS = (
    "number",
    "violation_type",
    "situation",
    "root_cause_description",
    "corrective_actions",
    "category",
    "factors",
    "impact_budget",
    "impact_schedule",
    "impact_quality",
    "proposals",
    "process_name",
    "author",
)


def parse_learned_lessons_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    sheet_name = _main_sheet_name(workbook.sheetnames)
    worksheet = workbook[sheet_name]

    metadata = _parse_metadata(worksheet)
    sections: list[dict[str, Any]] = []
    lessons: list[dict[str, Any]] = []
    current_section: dict[str, Any] | None = None

    for row in worksheet.iter_rows(min_row=11, values_only=True):
        cells = [_normalize_cell(value) for value in row[:13]]
        number = cells[0]
        if not number and not any(cells[1:]):
            continue

        first_cell = cells[0] or ""
        if first_cell.lower().startswith("примечания"):
            break

        if _is_reference_row(cells):
            continue

        if _is_section_row(cells):
            current_section = {
                "number": number,
                "title": cells[1],
            }
            sections.append(current_section)
            continue

        if not _is_lesson_row(cells):
            continue

        lesson = {
            "number": number,
            "section_number": current_section["number"] if current_section else None,
            "section_title": current_section["title"] if current_section else None,
            "violation_type": cells[1],
            "situation": cells[2],
            "root_cause_description": cells[3],
            "corrective_actions": cells[4],
            "category": cells[5],
            "factors": cells[6],
            "impact_budget": cells[7],
            "impact_schedule": cells[8],
            "impact_quality": cells[9],
            "proposals": cells[10],
            "process_name": cells[11],
            "author": cells[12],
        }
        lessons.append(lesson)

    workbook.close()
    return {
        "sheet_name": sheet_name,
        "metadata": metadata,
        "sections": sections,
        "lessons": lessons,
        "summary": {
            "sections_count": len(sections),
            "lessons_count": len(lessons),
        },
    }


def _main_sheet_name(sheet_names: list[str]) -> str:
    for name in sheet_names:
        lowered = name.lower()
        if "форм" in lowered or "сесси" in lowered:
            return name
    return sheet_names[0]


def _parse_metadata(worksheet: Any) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for row in worksheet.iter_rows(min_row=1, max_row=10, values_only=True):
        label = _normalize_cell(row[0] if row else "")
        if not label:
            continue
        normalized = label.lower().strip()
        for marker, key in METADATA_LABELS.items():
            if marker in normalized:
                value = _normalize_cell(row[2] if len(row) > 2 else "")
                if not value and len(row) > 3:
                    value = _normalize_cell(row[3])
                if value:
                    metadata[key] = value
                if normalized.startswith("должность") and len(row) > 3:
                    person = _normalize_cell(row[3])
                    if person:
                        metadata["responsible_person"] = person
    return metadata


def _is_lesson_row(cells: list[str]) -> bool:
    number = cells[0]
    if not LESSON_NUMBER_RE.match(number):
        return False
    if "." not in number and number.isdigit():
        return False
    content_columns = cells[1:11]
    return any(content_columns)


def _is_section_row(cells: list[str]) -> bool:
    number = cells[0]
    if not number or "." in number:
        return False
    if not number.isdigit():
        return False
    return bool(cells[1]) and not any(cells[2:5])


def _is_reference_row(cells: list[str]) -> bool:
    if cells[0]:
        return False
    return bool(cells[1]) and not any(cells[2:5])


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)
